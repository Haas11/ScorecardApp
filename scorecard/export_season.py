from __future__ import annotations

import re
import sqlite3
from pathlib import Path
from typing import Optional

import click
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import yaml

from db import get_connection, init_db
from stats import compute_all_stats, PlayerStats

_CONFIG_PATH = Path("config.yml")
_DB_PATH = Path("data/season.db")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TEAM_FILL = PatternFill("solid", fgColor="D9E1F2")
TEAM_FONT = Font(bold=True)

# #.000 → ".308" for <1, "1.583" for >=1; no leading zero on rate stats
_COL_FMT: dict[str, str] = {
    "AVG": "#.000", "OBP": "#.000", "SLG": "#.000", "OPS": "#.000",
    "BABIP": "#.000", "ISO": "#.000", "wOBA": "#.000",
    "BB%": "#.000", "K%": "#.000", "BB/K": "#.000",
    "RC": "0.00", "OPS+": "0", "AB/HR": "0.0",
}

RATE_COLS = {"AVG", "OBP", "SLG", "OPS", "BABIP", "ISO", "wOBA"}
SEASON_COLS = [
    "Name", "G", "PA", "AB", "H", "2B", "3B", "HR", "R", "RBI",
    "BB", "K", "SB", "CS", "AVG", "OBP", "SLG", "OPS",
    "BABIP", "ISO", "BB%", "K%", "wOBA", "RC", "OPS+", "AB/HR", "BB/K",
]
GAME_LOG_COLS = [
    "Name", "Date", "Opponent", "PA", "AB", "H", "2B", "3B", "HR",
    "R", "RBI", "BB", "K", "SB", "CS", "AVG", "OBP",
]
LOW_CONF_COLS = ["Player", "Date", "Opponent", "Inning", "Result", "Notes", "Reviewed"]


def _fmt(val, col: str):
    if val is None:
        return ""
    if col in RATE_COLS or col in ("BB%", "K%"):
        return round(float(val), 3)
    return val


def _write_header(ws, cols: list[str]) -> None:
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 8)


def _stats_row(s: PlayerStats) -> list:
    return [
        s.name, s.G, s.PA, s.AB, s.H, s.doubles, s.triples, s.HR,
        s.R, s.RBI, s.BB, s.K, s.SB, s.CS,
        round(s.AVG, 3), round(s.OBP, 3), round(s.SLG, 3), round(s.OPS, 3),
        round(s.BABIP, 3) if s.BABIP is not None else "",
        round(s.ISO, 3),
        round(s.BB_pct, 3), round(s.K_pct, 3),
        round(s.wOBA, 3) if s.wOBA is not None else "",
        round(s.RC, 2),
        round(s.OPS_plus, 1),
        round(s.AB_per_HR, 1) if s.AB_per_HR is not None else "",
        round(s.BB_K, 2) if s.BB_K is not None else "",
    ]


def _team_season_row(stats: list) -> list:
    if not stats:
        return ["Team"] + [""] * (len(SEASON_COLS) - 1)
    g = max(s.G for s in stats)
    pa = sum(s.PA for s in stats)
    ab = sum(s.AB for s in stats)
    h = sum(s.H for s in stats)
    d = sum(s.doubles for s in stats)
    t = sum(s.triples for s in stats)
    hr = sum(s.HR for s in stats)
    r = sum(s.R for s in stats)
    rbi = sum(s.RBI for s in stats)
    bb = sum(s.BB for s in stats)
    k = sum(s.K for s in stats)
    sb = sum(s.SB for s in stats)
    cs = sum(s.CS for s in stats)
    avg = round(h / ab, 3) if ab > 0 else 0.0
    obp = round((h + bb) / (ab + bb), 3) if (ab + bb) > 0 else 0.0
    singles = h - d - t - hr
    tb = singles + 2 * d + 3 * t + 4 * hr
    slg = round(tb / ab, 3) if ab > 0 else 0.0
    ops = round(obp + slg, 3)
    babip_denom = ab - k - hr
    babip = round((h - hr) / babip_denom, 3) if babip_denom > 0 else ""
    iso = round(slg - avg, 3)
    bb_pct = round(bb / pa, 3) if pa > 0 else 0.0
    k_pct = round(k / pa, 3) if pa > 0 else 0.0
    ab_hr = round(ab / hr, 1) if hr > 0 else ""
    bb_k = round(bb / k, 2) if k > 0 else ""
    return [
        "Team", g, pa, ab, h, d, t, hr, r, rbi, bb, k, sb, cs,
        avg, obp, slg, ops, babip, iso, bb_pct, k_pct, "", "", "", ab_hr, bb_k,
    ]


def _team_game_row(entries: list[dict]) -> list:
    pa = sum(d["PA"] for d in entries)
    ab = sum(d["AB"] for d in entries)
    h = sum(d["H"] for d in entries)
    d2 = sum(d["2B"] for d in entries)
    t = sum(d["3B"] for d in entries)
    hr = sum(d["HR"] for d in entries)
    r = sum(d["R"] for d in entries)
    rbi = sum(d["RBI"] for d in entries)
    bb = sum(d["BB"] for d in entries)
    k = sum(d["K"] for d in entries)
    sb = sum(d["SB"] for d in entries)
    cs = sum(d["CS"] for d in entries)
    avg = round(h / ab, 3) if ab > 0 else 0.0
    obp = round((h + bb) / (ab + bb), 3) if (ab + bb) > 0 else 0.0
    return ["", "Team", pa, ab, h, d2, t, hr, r, rbi, bb, k, sb, cs, avg, obp]


def _write_team_row(ws, row_idx: int, vals: list, cols: list[str] | None = None) -> None:
    for col_idx, val in enumerate(vals, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = TEAM_FONT
        cell.fill = TEAM_FILL
        if cols and col_idx <= len(cols) and cols[col_idx - 1] in _COL_FMT:
            cell.number_format = _COL_FMT[cols[col_idx - 1]]


def _add_color_scale(
    ws, col_idx: int, min_val: float, mid_val: float, max_val: float,
    num_rows: int, start_row: int = 2,
) -> None:
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}{start_row}:{col_letter}{start_row + num_rows - 1}"
    rule = ColorScaleRule(
        start_type="num", start_value=min_val, start_color="FFFFFF",   # white
        mid_type="num",   mid_value=mid_val,   mid_color="FFEB9C",     # light yellow
        end_type="num",   end_value=max_val,   end_color="63BE7B",     # green
    )
    ws.conditional_formatting.add(cell_range, rule)


def export_season(
    output_path: str = "stats.xlsx",
    min_pa: int = 0,
    db_path: Path | None = None,
) -> None:
    conn = get_connection(db_path or _DB_PATH)
    init_db(db_path or _DB_PATH)

    cfg = yaml.safe_load(open(_CONFIG_PATH))
    lg_obp = cfg["league"]["avg_obp"]
    lg_ops = cfg["league"]["avg_ops"]

    all_stats = compute_all_stats(conn)
    filtered = [s for s in all_stats if s.PA >= min_pa]
    filtered.sort(key=lambda s: s.OPS, reverse=True)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Season Stats ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Season Stats"
    _write_header(ws1, SEASON_COLS)

    _write_team_row(ws1, 2, _team_season_row(all_stats), SEASON_COLS)

    italic_font = Font(italic=True)
    for row_idx, s in enumerate(filtered, 3):
        row_data = _stats_row(s)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            col_name = SEASON_COLS[col_idx - 1]
            if col_name in _COL_FMT:
                cell.number_format = _COL_FMT[col_name]
            if s.small_sample:
                cell.font = italic_font

    # Color scale on AVG, OBP, OPS, wOBA — (floor, league-avg, excellent)
    color_scale_targets = {
        "AVG":  (0.150, 0.260, 0.400),
        "OBP":  (0.200, lg_obp, 0.450),
        "OPS":  (0.450, lg_ops, 1.100),
        "wOBA": (0.200, 0.320, 0.430),
    }
    num_rows = len(filtered)
    for col_name, (min_v, mid_v, max_v) in color_scale_targets.items():
        if col_name in SEASON_COLS:
            col_idx = SEASON_COLS.index(col_name) + 1
            _add_color_scale(ws1, col_idx, min_v, mid_v, max_v, num_rows, start_row=3)

    _autofit(ws1)

    # ── Sheet 2: Game Log ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Game Log")
    _write_header(ws2, GAME_LOG_COLS)

    rows = conn.execute(
        """SELECT p.name, g.date, g.opponent,
                  pa.batting_order, pa.inning,
                  pa.result, pa.run_scored, pa.rbi, pa.bb, pa.hp,
                  pa.sac, pa.sf, pa.sb, pa.cs, pa.game_id, p.player_id
           FROM plate_appearances pa
           JOIN players p ON pa.player_id = p.player_id
           JOIN games g ON pa.game_id = g.game_id
           ORDER BY g.date ASC, pa.batting_order ASC""",
    ).fetchall()

    # Aggregate per player per game
    from collections import defaultdict

    game_player: dict[tuple, dict] = defaultdict(lambda: {
        "PA": 0, "AB": 0, "H": 0, "2B": 0, "3B": 0, "HR": 0,
        "R": 0, "RBI": 0, "BB": 0, "K": 0, "SB": 0, "CS": 0,
        "name": "", "date": "", "opponent": "", "batting_order": 99,
    })

    for r in rows:
        key = (r["player_id"], r["game_id"])
        d = game_player[key]
        d["name"] = r["name"]
        d["date"] = r["date"] or ""
        d["opponent"] = r["opponent"] or ""
        d["batting_order"] = min(d["batting_order"], r["batting_order"])
        d["PA"] += 1
        result = (r["result"] or "").upper()
        bb = r["bb"]
        hp = r["hp"]
        sac = r["sac"]
        sf = r["sf"]
        if not (bb or hp or sac or sf):
            d["AB"] += 1
        if result in ("1B", "2B", "3B", "HR"):
            d["H"] += 1
        if result == "2B":
            d["2B"] += 1
        if result == "3B":
            d["3B"] += 1
        if result == "HR":
            d["HR"] += 1
        d["R"] += r["run_scored"]
        d["RBI"] += r["rbi"]
        d["BB"] += bb
        if result in ("K", "KL"):
            d["K"] += 1
        d["SB"] += r["sb"]
        d["CS"] += r["cs"]

    game_log_rows = sorted(
        game_player.values(),
        key=lambda x: (x["date"], x["batting_order"]),
    )
    for row_idx, d in enumerate(game_log_rows, 2):
        ab = d["AB"]
        h = d["H"]
        bb = d["BB"]
        hp_val = 0
        sf_val = 0
        avg = round(h / ab, 3) if ab > 0 else 0.0
        obp_denom = ab + bb + hp_val + sf_val
        obp = round((h + bb + hp_val) / obp_denom, 3) if obp_denom > 0 else 0.0
        row_vals = [
            d["name"], d["date"], d["opponent"],
            d["PA"], d["AB"], d["H"], d["2B"], d["3B"], d["HR"],
            d["R"], d["RBI"], d["BB"], d["K"], d["SB"], d["CS"],
            avg, obp,
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            col_name = GAME_LOG_COLS[col_idx - 1]
            if col_name in _COL_FMT:
                cell.number_format = _COL_FMT[col_name]

    _autofit(ws2)

    # ── Per-game sheets: one box-score tab per game (regenerated from DB) ──
    per_game_cols = ["#", "Name", "PA", "AB", "H", "2B", "3B", "HR",
                     "R", "RBI", "BB", "K", "SB", "CS", "AVG", "OBP"]
    games_meta = conn.execute(
        "SELECT game_id, date, opponent, game_number FROM games ORDER BY date ASC, game_id ASC"
    ).fetchall()
    used_names: set[str] = set()
    for gm in games_meta:
        entries = [d for (pid, gid), d in game_player.items() if gid == gm["game_id"]]
        if not entries:
            continue
        entries.sort(key=lambda x: x["batting_order"])
        base = f"{gm['date'] or '?'} {gm['opponent'] or '?'}"
        if gm["game_number"]:
            base += f" g{gm['game_number']}"
        sheet_name = re.sub(r"[\[\]:*?/\\]", "", base)[:31] or f"game{gm['game_id']}"
        unique, k = sheet_name, 2
        while unique in used_names:
            unique = (sheet_name[:28] + f"_{k}")[:31]
            k += 1
        used_names.add(unique)

        wsg = wb.create_sheet(unique)
        _write_header(wsg, per_game_cols)
        _write_team_row(wsg, 2, _team_game_row(entries), per_game_cols)
        for row_idx, d in enumerate(entries, 3):
            ab, h, bb = d["AB"], d["H"], d["BB"]
            avg = round(h / ab, 3) if ab > 0 else 0.0
            obp = round((h + bb) / (ab + bb), 3) if (ab + bb) > 0 else 0.0
            vals = [d["batting_order"], d["name"], d["PA"], d["AB"], d["H"],
                    d["2B"], d["3B"], d["HR"], d["R"], d["RBI"], d["BB"],
                    d["K"], d["SB"], d["CS"], avg, obp]
            for col_idx, val in enumerate(vals, 1):
                cell = wsg.cell(row=row_idx, column=col_idx, value=val)
                col_name = per_game_cols[col_idx - 1]
                if col_name in _COL_FMT:
                    cell.number_format = _COL_FMT[col_name]
        _autofit(wsg)

    # ── Sheet 3: Low Confidence ────────────────────────────────────────
    ws3 = wb.create_sheet("Low Confidence")
    _write_header(ws3, LOW_CONF_COLS)

    lc_rows = conn.execute(
        """SELECT p.name, g.date, g.opponent, pa.inning,
                  pa.result, pa.raw_notes, pa.reviewed
           FROM plate_appearances pa
           JOIN players p ON pa.player_id = p.player_id
           JOIN games g ON pa.game_id = g.game_id
           WHERE pa.needs_review = 1
           ORDER BY g.date ASC, p.name ASC""",
    ).fetchall()

    for row_idx, r in enumerate(lc_rows, 2):
        ws3.cell(row=row_idx, column=1, value=r["name"])
        ws3.cell(row=row_idx, column=2, value=r["date"])
        ws3.cell(row=row_idx, column=3, value=r["opponent"])
        ws3.cell(row=row_idx, column=4, value=r["inning"])
        ws3.cell(row=row_idx, column=5, value=r["result"])
        ws3.cell(row=row_idx, column=6, value=r["raw_notes"])
        ws3.cell(row=row_idx, column=7, value="Yes" if r["reviewed"] else "No")

    _autofit(ws3)

    wb.save(output_path)
    print(f"Exported stats to {output_path}")
    conn.close()


@click.command()
@click.option("--output", default="stats.xlsx", help="Output Excel file path")
@click.option("--min-pa", default=0, type=int, help="Minimum PA to include player")
def main(output: str, min_pa: int) -> None:
    export_season(output, min_pa)


if __name__ == "__main__":
    main()
